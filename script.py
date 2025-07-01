import sys
import os
import subprocess
import datetime
import traceback

# Настройка кодировки по умолчанию
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

def emergency_log(message):
    """Экстренное логирование в консоль и файл при критических сбоях"""
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f"{timestamp} - {message}")  # Всегда выводим в консоль
    
    try:
        with open('logs.txt', 'a', encoding='utf-8', errors='replace') as f:
            f.write(f"{timestamp} - {message}\n")
    except Exception as e:
        print(f"{timestamp} - Невозможно записать лог: {str(e)}")

def log_message(message):
    """Основная функция логирования"""
    try:
        timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open('logs.txt', 'a', encoding='utf-8', errors='replace') as f:
            f.write(f"{timestamp} - {message}\n")
        print(f"{timestamp} - {message}")  # Дублируем в консоль
    except Exception as e:
        emergency_log(f"Сбой логирования: {str(e)}. Исходное сообщение: {message}")

def install_library(library):
    """Установка библиотек с улучшенным логированием"""
    try:
        log_message(f"Попытка установки {library}")
        result = subprocess.run(
            [sys.executable, '-m', 'pip', 'install', '--user', library],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace'
        )
        
        if result.returncode == 0:
            log_message(f"{library} успешно установлена")
            return True
        else:
            log_message(f"Ошибка установки {library}:\n"
                       f"STDOUT: {result.stdout}\n"
                       f"STDERR: {result.stderr}")
            return False
    except Exception as e:
        log_message(f"Критическая ошибка при установке {library}: {str(e)}")
        return False

def main():
    try:
        # Экстренное логирование старта
        emergency_log("Начало выполнения скрипта")
        
        # Проверка версии Python
        if sys.version_info < (3, 6):
            msg = "Требуется Python 3.6 или новее!"
            emergency_log(msg)
            raise RuntimeError(msg)

        # Проверка рабочей директории
        emergency_log(f"Текущая директория: {os.getcwd()}")
        emergency_log(f"Файлы в директории: {os.listdir()}")
        
        # Проверка и установка зависимостей
        required = ['openpyxl']
        for lib in required:
            try:
                __import__(lib)
                log_message(f"{lib} уже установлена")
            except ImportError:
                if not install_library(lib):
                    raise RuntimeError(f"Не удалось установить {lib}")
                try:
                    __import__(lib)
                except ImportError:
                    raise RuntimeError(f"Ошибка импорта после установки {lib}")

        # Основная логика
        from openpyxl import load_workbook
        
        if not os.path.exists('example.xlsx'):
            raise FileNotFoundError("Файл example.xlsx не найден")

        wb = load_workbook(filename='example.xlsx', read_only=True)
        ws = wb.active
        
        output = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            cleaned = [str(cell).replace('\n', ' ').strip() if cell is not None else '' for cell in row]
            output.append(' '.join(cleaned))
        
        with open('data.txt', 'w', encoding='utf-8', errors='replace') as f:
            f.write('\n'.join(output))
        
        log_message(f"Успешно записано {len(output)} строк")

    except Exception as e:
        error_msg = f"КРИТИЧЕСКАЯ ОШИБКА: {str(e)}\n{traceback.format_exc()}"
        emergency_log(error_msg)
    finally:
        if 'wb' in locals():
            wb.close()
        emergency_log("Завершение работы скрипта")

if __name__ == "__main__":
    main()